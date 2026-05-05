import pandas as pd
import numpy as np
import datetime
from datetime import *
import warnings
from cbhpacks.bins_model import *
from cbhpacks.cols_select import *
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from lightgbm import LGBMClassifier
from xgboost import XGBClassifier
from datetime import datetime,timedelta
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import chi2
from sklearn.datasets import make_regression, make_classification
from sklearn.preprocessing import MinMaxScaler
from sklearn.feature_selection import RFE
from sklearn.feature_selection import SelectFromModel
from sklearn.metrics import roc_curve
from sklearn.metrics import roc_auc_score, auc
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.model_selection import train_test_split
warnings.filterwarnings("ignore")
from sklearn.model_selection import GridSearchCV
import joblib
import seaborn as sbn
import matplotlib.pyplot as plt
from skopt import BayesSearchCV #安装时用pip install scikit-optimize
import os
from sklearn.decomposition import PCA
import jieba
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import KMeans
from collections import Counter
import matplotlib.pyplot as plt
from sklearn import metrics
from sklearn.metrics import silhouette_score
from sklearn.cluster import DBSCAN
from scipy.cluster.hierarchy import linkage, fcluster, dendrogram
import statsmodels.api as sm
import statsmodels.formula.api as smf
from keras.models import Sequential
from keras.layers import Dense
from sklearn.svm import SVC
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule,ColorScaleRule
from openpyxl.styles import Color,PatternFill, Font
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.drawing.image import Image
'''
包含二分类模型binary_model、无监督模型uns_model、线性回归显著性检验模型linear_model
'''

class binary_model:
    '''
    用法：
    用lr_fit或xgb_fit或lgbm_fit训练，用report自动输出对应训练模型的报告
    mt=binary_model(df=woe_df,cols=cols,target='target',shuffle=True,random_state=666,test_size=0.2,model_path='step6_binary_model_training' ,train_data_path='step6_binary_model_training/datas')
    mt.lr_fit(penalty='l2',C=1,solver='saga',max_iter=100)
    mt.xgb_fit(colsample_bytree=None,learning_rate=None,num_leaves=None,min_child_weight=None
                    ,min_split_gain=None ,reg_alpha=None,reg_lambda=None,max_depth=None,n_estimators=100)
    mt.lgbm_fit(colsample_bytree=0.35,learning_rate=0.14,num_leaves=13,min_child_weight=45
                    ,min_gain_to_split=0.45 ,reg_alpha=2.2,reg_lambda=2.2,max_depth=5,n_estimators=110,metric='roc_auc')
    mt.mlp_fit(self,epochs=30,batch_size=10,validation_split=0.2,loss='binary_crossentropy',optimizer='adam',metrics=['AUC','recall'])
    bins_rpt,confusion_matrix=mt.report(group=30,bins_type='eq_cnt')
    para_dic=mt.para_adj_gs(paras={'learning_rate':[0.05,0.1,0.2],'colsample_bytree':[0.6,0.8,1],'num_leaves':[24,28,31]},score_type='roc_auc',cv=2)
    para_dic=mt.para_adj_bs(paras={'learning_rate':[0.05,0.1,0.2],'colsample_bytree':[0.6,0.8,1],'num_leaves':[24,28,31]},score_type='roc_auc',cv=2)
    每次训练模型都会更新self.clf

    随着建模过程，会多两个参数，self.clf和self.model_type，clf是训练使用的模型，model_type是模型名称简称（xgb,lgmb,lr）
    '''
    
    def __init__(self,train,test,cols,target,model_path='step6_binary_model' ,train_data_path='step6_binary_model/datas',save=False):
        self.xtrain=train
        self.xtest=test
        self.cols=cols
        self.target=target
        self.cols_copy=cols.copy()
        self.adj='noadj'
        self.df=pd.concat([train,test]).reset_index(drop=True)
        self.cols_bins_rpt=''
        
        self.model_path=model_path
        self.train_data_path=train_data_path
        dir_name_list=[model_path,train_data_path]
        current_path = os.getcwd()
        for dir_name in dir_name_list:
            path=os.path.join(os.getcwd(), dir_name)
            try:
                os.mkdir(path)  # 使用 os.makedirs(path, exist_ok=True) 可以避免多级目录的创建
                print(f"文件夹 '{dir_name}' 已成功创建在 '{current_path}'.")
            except FileExistsError:
                print(f"文件夹 '{dir_name}' 已存在于 '{current_path}'.")
            except PermissionError:
                print(f"无法创建文件夹于该路径 '{current_path}'.")
        if save:
            try:
                self.xtrain[self.cols].to_csv(self.train_data_path+'/train.csv',index=None)
                self.xtest[self.cols].to_csv(self.train_data_path+'/test.csv',index=None)
            except:
                print('未设定好训练集测试集，可用class.train_test_split()函数设定')
        else:
            pass

    def train_test_split(self,df,shuffle,random_state,test_size):
        self.df=df
        self.xtrain,self.xtest,self.ytrain,self.ytest=train_test_split(df,df[self.target],shuffle=shuffle,random_state=random_state,test_size=test_size)
        self.xtrain[self.cols].to_csv(self.train_data_path+'/train.csv',index=None)
        self.xtest[self.cols].to_csv(self.train_data_path+'/test.csv',index=None)

        
    def lr_fit(self,*args, **kwargs):
        '''
penalty:选择正则化类型：penalty 参数可以是 l1、l2、elasticnet 或 none。选择合适的正则化类型可以帮助防止模型过拟合。
        l1 正则化倾向于生成稀疏的模型权重，即使得某些特征的系数为零。
        l2 正则化倾向于使得所有特征的权重均较小，但不会为零。
        elasticnet 是 l1 和 l2 的组合，通过调整 l1_ratio 参数来平衡两者的影响。默认l2
C：调整正则化强度：通过 C 参数（正则化的倒数）来控制强度。较小的 C 值意味着更强的正则化。默认1

solver:liblinear 适用于小数据集，支持 l1 和 l2 正则化。
        lbfgs、newton-cg、sag 和 saga 支持大数据集和多分类问题。sag 和 saga 是针对大数据集特别优化的。默认saga
max_iter：默认100，算法求解最大迭代次数，默认是100。
        '''
        print("---------------------------------------------")
        lr=LogisticRegression(*args, **kwargs)
        lr.fit(self.xtrain[self.cols],self.xtrain[self.target])
        print("类别:", lr.classes_)
        print("截距:", lr.intercept_)
        for idx, coef in enumerate(lr.coef_[0]):
            print(f"特征 {self.cols[idx]}的系数为: {coef:.4f}")
        self.clf=lr
        self.model_type='lr'
        lr_coef_df=pd.DataFrame()
        lr_coef_df['col_name']=self.cols
        lr_coef_df['coef']=lr.coef_[0]
        lr_coef_df.to_excel(self.model_path+'/lr_coef.xlsx',index=None)
        self.cols_weight=lr_coef_df
        joblib.dump(lr,self.model_path+'/lr_model.pkl')
        self.adj='noadj'
        print("---------------------------------------------")

    def xgb_fit(self,*args, **kwargs):
        print("---------------------------------------------")
        xgb=XGBClassifier(*args, **kwargs)
        xgb.fit(self.xtrain[self.cols],self.xtrain[self.target])
        importances_data=pd.DataFrame()
        importances_data['col_name']=self.cols
        importances_data['importances']=xgb.feature_importances_
        for idx, imp in enumerate(importances_data.importances):
            print(f"特征 {importances_data['col_name'][idx]}的重要性为: {imp:.4f}")
        self.clf=xgb
        self.model_type='xgb'
        importances_data.to_excel(self.model_path+'/xgb_imp.xlsx',index=None)
        joblib.dump(xgb,self.model_path+'/xgb_model.pkl')
        self.cols_weight=importances_data
        self.adj='noadj'
        print("---------------------------------------------")

    def lgbm_fit(self,*args, **kwargs):
        print("---------------------------------------------")
        lgbm=LGBMClassifier(*args, **kwargs)
        lgbm.fit(self.xtrain[self.cols],self.xtrain[self.target])
        importances_data=pd.DataFrame()
        importances_data['col_name']=self.cols
        importances_data['importances']=lgbm.feature_importances_
        for idx, imp in enumerate(importances_data.importances):
            print(f"特征 {importances_data['col_name'][idx]}的重要性为: {imp:.4f}")
        self.clf=lgbm
        self.model_type='lgbm'
        importances_data.to_excel(self.model_path+'/lgbm_imp.xlsx',index=None)
        joblib.dump(lgbm,self.model_path+'/lgbm_model.pkl')
        self.cols_weight=importances_data
        self.adj='noadj'
        print("---------------------------------------------")


    def mlp_fit(self,epochs=5,batch_size=100,validation_split=0.5,metrics=[],loss='binary_crossentropy',optimizer='adam'):
        print("---------------------------------------------")
        mlp = Sequential()
        mlp.add(Dense(64, activation='relu', input_dim=self.xtrain[self.cols].shape[1]))
        mlp.add(Dense(32, activation='relu'))
        mlp.add(Dense(1, activation='sigmoid'))  # 输出层
        mlp.compile(loss=loss, optimizer=optimizer, metrics=metrics)
        mlp.fit(self.xtrain[self.cols], self.xtrain[self.target], epochs=epochs, batch_size=batch_size, validation_split=validation_split)
        self.clf=mlp
        self.model_type='keras'
        weight_data=pd.DataFrame({'col_name':self.cols,'weight':np.abs(self.clf.get_weights()[0]).sum(axis=1)})
        self.cols_weight=weight_data
        #joblib.dump(mlp.get_weights(),self.model_path+'/mlp_weights.h5')
        weight_data.to_csv(self.model_path+'/mlp_weight.csv',index=None)
        mlp.save(self.model_path+'/mlp_model.h5')
        self.adj='noadj'
        print("---------------------------------------------")

    def svm_fit(self,*args, **kwargs):
        print("---------------------------------------------")
        svm = SVC(kernel='linear',probability=True,*args, **kwargs)  # 你可以选择其他内核，如'poly', 'rbf', 'sigmoid'等
        svm.fit(self.xtrain[self.cols],self.xtrain[self.target])
        importances_data=pd.DataFrame()
        importances_data['col_name']=self.cols
        importances_data['importances']=svm.coef_[0]
        for idx, imp in enumerate(importances_data.importances):
            print(f"特征 {importances_data['col_name'][idx]}的重要性为: {imp:.4f}")
        self.clf=svm
        self.model_type='svm'
        importances_data.to_excel(self.model_path+'/svm_imp.xlsx',index=None)
        joblib.dump(svm,self.model_path+'/svm_model.pkl')
        self.cols_weight=importances_data
        self.adj='noadj'
        print("---------------------------------------------")

    def rdf_fit(self,*args, **kwargs):
        print("---------------------------------------------")
        rdf=RandomForestClassifier(*args, **kwargs)
        rdf.fit(self.xtrain[self.cols],self.xtrain[self.target])
        importances_data=pd.DataFrame()
        importances_data['col_name']=self.cols
        importances_data['importances']=rdf.feature_importances_
        for idx, imp in enumerate(importances_data.importances):
            print(f"特征 {importances_data['col_name'][idx]}的重要性为: {imp:.4f}")
        self.clf=rdf
        self.model_type='rdf'
        importances_data.to_excel(self.model_path+'/rdf_imp.xlsx',index=None)
        joblib.dump(rdf,self.model_path+'/rdf_model.pkl')
        self.cols_weight=importances_data
        self.adj='noadj'
        print("---------------------------------------------")


    def para_adj_gs(self,paras,score_type='roc_auc',cv=2):
        print("---------------------------------------------")
        gs=GridSearchCV(self.clf,paras,scoring=score_type,cv=cv)
        gs.fit(self.xtrain[self.cols],self.xtrain[self.target])
        print(gs.best_params_)
        joblib.dump(gs,self.model_path+'/'+self.model_type+'_gs_adj.pkl')
        para_dic=gs.best_params_
        self.clf.set_params(**para_dic)
        self.clf.fit(self.xtrain[self.cols],self.xtrain[self.target])
        joblib.dump(self.clf,self.model_path+'/'+self.model_type+'_gs_adj_model_final.pkl')
        self.adj='gs_adj'
        print("---------------------------------------------")
        return para_dic

    def para_adj_bs(self,paras,score_type='roc_auc',cv=2):
        print("---------------------------------------------")
        bs=BayesSearchCV(self.clf,paras,scoring=score_type,cv=cv, n_jobs=-1, verbose=-1)
        bs.fit(self.xtrain[self.cols],self.xtrain[self.target])
        print(dict(bs.best_params_))
        joblib.dump(bs,self.model_path+'/'+self.model_type+'_bs_adj.pkl')
        para_dic=dict(bs.best_params_)
        self.clf.set_params(**para_dic)
        self.clf.fit(self.xtrain[self.cols],self.xtrain[self.target])
        joblib.dump(self.clf,self.model_path+'/'+self.model_type+'_bs_adj_model_final.pkl')
        self.adj='bs_adj'
        print("---------------------------------------------")
        return para_dic

    def report(self,group,mth_col,base_mth,bins_type='all',cols_bins_rpt=''):
        '''
        group:评分分数分组个数int
        bins_type：评分分数分组类型，默认为all（等频+等距），也支持eq_cnt或eq_distance
        mth_col：月份字段名称str
        map：如果按照分箱建模，将已有的WOE分箱细则dict作为分箱模板分箱。
        '''
        print("---------------------------------------------")
        print("正在产出混淆矩阵及评估指标...")
        self.cols_bins_rpt=cols_bins_rpt
        def card_score(pred,P=660,Q=20,PDO=50,good_weight=1):
            Odds = (1-pred)*good_weight/pred 
            B = -PDO/np.log(Q)
            A = P + B*np.log(Q)
            return A - B*np.log(Odds)
        if self.model_type=='keras':
            y_pred_prob_a=self.clf.predict(self.xtest[self.cols])
            y_pred_prob_a_train=self.clf.predict(self.xtrain[self.cols])
            y_pred_prob_a_all=self.clf.predict(self.df[self.cols])
            y_pred_prob_b=1-y_pred_prob_a
            y_pred_prob_b_train=1-y_pred_prob_a_train
            y_pred_prob_b_all=1-y_pred_prob_a_all
            y_pred_prob_train=np.hstack((y_pred_prob_b_train, y_pred_prob_a_train))
            y_pred_prob_all=np.hstack((y_pred_prob_b_all, y_pred_prob_a_all))
            y_pred_prob=np.hstack((y_pred_prob_b, y_pred_prob_a))
            y_pred_train = (y_pred_prob_train[:,1] > 0.5).astype(int)
            y_pred_all = (y_pred_prob_all[:,1] > 0.5).astype(int)
            y_pred = (y_pred_prob[:,1] > 0.5).astype(int) 
        else:
            y_pred_prob_train=self.clf.predict_proba(self.xtrain[self.cols])
            y_pred_prob_all=self.clf.predict_proba(self.df[self.cols])
            y_pred_prob=self.clf.predict_proba(self.xtest[self.cols])
            y_pred_train=self.clf.predict(self.xtrain[self.cols])
            y_pred_all=self.clf.predict(self.df[self.cols])
            y_pred=self.clf.predict(self.xtest[self.cols])
        # 计算模型总体效果
        accuracy = accuracy_score(self.xtest[self.target], y_pred)
        accuracy_train = accuracy_score(self.xtrain[self.target], y_pred_train)
        accuracy_all = accuracy_score(self.df[self.target], y_pred_all)
        precision = precision_score(self.xtest[self.target], y_pred)
        precision_train = precision_score(self.xtrain[self.target], y_pred_train)
        precision_all = precision_score(self.df[self.target], y_pred_all)
        recall = recall_score(self.xtest[self.target], y_pred)
        recall_train = recall_score(self.xtrain[self.target], y_pred_train)
        recall_all = recall_score(self.df[self.target], y_pred_all)
        f1 = f1_score(self.xtest[self.target], y_pred)
        f1_train = f1_score(self.xtrain[self.target], y_pred_train)
        f1_all = f1_score(self.df[self.target], y_pred_all)
        auc_score=roc_auc_score(self.xtest[self.target],y_pred_prob[:,1])
        auc_score_train=roc_auc_score(self.xtrain[self.target],y_pred_prob_train[:,1])
        auc_score_all=roc_auc_score(self.df[self.target],y_pred_prob_all[:,1])
        fpr,tpr,thres=roc_curve(self.xtest[self.target],y_pred_prob[:,1])
        fpr_train,tpr_train,thres_train=roc_curve(self.xtrain[self.target],y_pred_prob_train[:,1])
        fpr_all,tpr_all,thres_all=roc_curve(self.df[self.target],y_pred_prob_all[:,1])
        ks=max(tpr-fpr)
        ks_train=max(tpr_train-fpr_train)
        ks_all=max(tpr_all-fpr_all)
        print(f"准确率————test:{accuracy:.4f}，train:{accuracy_train:.4f}，all:{accuracy_all:.4f}")
        print(f"精确率————test:{precision:.4f}，train:{precision_train:.4f}，all:{precision_all:.4f}")
        print(f"召回率————test:{recall:.4f}，train:{recall_train:.4f}，all:{recall_all:.4f}")
        print(f"F1分数————test:{f1:.4f}，train:{f1_train:.4f}，all:{f1_all:.4f}")
        print(f"KS分数————test:{ks:.4f}，train:{ks_train:.4f}，all:{ks_all:.4f}")
        print(f"ROCAUC————test:{auc_score:.4f}，train:{auc_score_train:.4f}，all:{auc_score_all:.4f}")
        confusion_matrix=pd.DataFrame([['test',accuracy,precision,recall,f1,auc_score,ks]
                                      ,['train',accuracy_train,precision_train,recall_train,f1_train,auc_score_train,ks_train]
                                      ,['all',accuracy_all,precision_all,recall_all,f1,auc_score_all,ks_all]]
                                      ,columns=['type','accuracy','precision','recall','F1','auc','ks'])
        confusion_matrix.to_excel(self.model_path+'/confusion_matrix_'+self.model_type+'_'+self.adj+'.xlsx',index=None)
        
        #计算模型分数、概率
        print("正在计算模型分数及概率...")
        self.xtest['y_pred_prob']=y_pred_prob[:,1]
        self.xtrain['y_pred_prob']=y_pred_prob_train[:,1]
        self.df['y_pred_prob']=y_pred_prob_all[:,1]
        self.xtest['y_pred']=y_pred
        self.xtrain['y_pred']=y_pred_train
        self.df['y_pred']=y_pred_all
        self.xtest['score']=card_score(y_pred_prob[:,1])
        self.xtrain['score']=card_score(y_pred_prob_train[:,1])
        self.df['score']=card_score(y_pred_prob_all[:,1])
        self.xtest['score'][self.xtest['score']<0]=0
        self.xtrain['score'][self.xtrain['score']<0]=0
        self.df['score'][self.df['score']<0]=0
        self.xtest[['y_pred_prob','y_pred','score']].to_csv(self.train_data_path+'/xtest_pred.csv')
        self.xtrain[['y_pred_prob','y_pred','score']].to_csv(self.train_data_path+'/xtrain_pred.csv')
        self.df[['y_pred_prob','y_pred','score']].to_csv(self.train_data_path+'/all_pred.csv')
        #计算模型分箱效果
        print("正在计算模型分数分箱效果及逐月效果...")
        if bins_type=='all':
            bt_list=['eq_cnt','eq_distance','cat_bin']
        elif bins_type=='eq_cnt':
            bt_list=['eq_cnt','cat_bin']
        elif bins_type=='eq_distance':
            bt_list=['eq_distance','cat_bin']
        data_type_list=[self.xtest,self.xtrain,self.df]
        bins_rpt_all=pd.DataFrame()
        for bt in bt_list:
            for datatype in data_type_list:  
                matrix_df=pd.DataFrame()
                bm=bins_model(df=datatype,cols=None,group=group,target=self.target,nan=-999999,bins_type=bt,mth_col=None,base_mth=None,cmp_mth=None,col='y_pred_prob'
                              ,adj_bin=False,min_group=1,bad_rate_adj=0,good_rate_adj=0,iv_adj=0,cat_cols=False,path=self.model_path)
                if bm.bins_type=='eq_cnt':
                    bin_edge,bins,bins_cnt=bm.eq_cnt()
                    datatype['bucket']=bins
                elif bm.bins_type=='eq_distance':
                    bin_edge,bins,bins_cnt=bm.eq_distance()
                    datatype['bucket']=bins
                elif bm.bins_type=='cat_bin':
                    bm.col=mth_col
                    bin_edge,bins,bins_cnt=bm.cat_bin()
                    datatype['bucket']=bins
                for key, grp in datatype.groupby('bucket'):
                    if datatype.equals(self.xtest):
                        grp['data_type']='test'
                    elif datatype.equals(self.xtrain):
                        grp['data_type']='train'
                    elif datatype.equals(self.df):
                        grp['data_type']='all'
                    grp['pred_bad']=grp.y_pred.sum()
                    try:
                        grp['accuracy']=len(grp[grp[self.target]==grp.y_pred])/len(grp)
                    except:
                        grp['accuracy']=np.nan
                    try:
                        grp['precision']=grp[grp.y_pred==1][self.target].sum()/len(grp[grp.y_pred==1])
                    except:
                        grp['precision']=np.nan
                    try:
                        grp['recall']=grp[grp[self.target]==1].y_pred.sum()/len(grp[grp[self.target]==1])
                    except:
                        grp['recall']=np.nan
                    try:
                        grp['f1']=(2*precision*recall)/(precision+recall)
                    except:
                        grp['f1']=np.nan
                    try:
                        auc_score_grp=roc_auc_score(grp[self.target],grp.y_pred_prob)
                        fpr_grp,tpr_grp,thres_grp=roc_curve(grp[self.target],grp.y_pred_prob)
                        ks_grp=max(tpr_grp-fpr_grp)
                        grp['ks_grp']=ks_grp#分箱再分箱后算出的分箱整体ks
                        grp['auc_grp']=auc_score_grp
                    except:
                        grp['ks_grp']=np.nan
                        grp['auc_grp']=np.nan
                    grp=grp[['bucket','accuracy','precision','recall','f1','ks_grp','auc_grp','pred_bad','data_type']].drop_duplicates()
                    matrix_df=pd.concat([matrix_df,grp])
                bins_rpt,iv=bm.bins_rpt()
                bins_rpt=pd.merge(left=bins_rpt,right=matrix_df,on='bucket')
                if datatype.equals(self.xtest):
                    bins_rpt['ks']=ks
                    bins_rpt['auc']=auc_score
                elif datatype.equals(self.xtrain):
                    bins_rpt['ks']=ks_train
                    bins_rpt['auc']=auc_score_train
                elif datatype.equals(self.df):
                    bins_rpt['ks']=ks_all
                    bins_rpt['auc']=auc_score_all
                bins_rpt=bins_rpt[['data_type','bins_type','bucket', 'col_name', 'total_cnt', 'bad_cnt', 'good_cnt', 'pred_bad','bad_rate','good_rate','badattr','goodattr','badattr_cum','goodattr_cum','accuracy','precision','recall','woe','f1', 'ks_bin','ks_grp','auc_grp', 'lift', 'ks', 'auc']]
                bins_rpt_all=pd.concat([bins_rpt_all,bins_rpt])
        bins_rpt_all.to_excel(self.model_path+'/bins_rpt_'+bins_type+'_'+self.model_type+'_'+self.adj+f'_{str(group)}_bins.xlsx',index=None)
        
        #计算入模特征效果
        #计算入模特征分箱
        print("正在计算入模特征分箱效果...")
        if len(self.cols_bins_rpt)==0:
            fea_bin=bins_model(df=self.df,cols=self.cols,group=5,target=self.target,nan=-999999,bins_type='eq_cnt',mth_col=mth_col,base_mth=base_mth,cmp_mth=None,col=None,adj_bin=False,cat_cols=False,path=self.model_path+'/bins_report')
            fea_bins_report,fea_iv=fea_bin.comp_woe_iv()
            psi_data=fea_bin.psi_mth_avg()
            self.cols_bins_rpt=fea_bins_report
        else:
            fea_bin=bins_model(df=self.df,cols=self.cols,group=5,target=self.target,nan=-999999,bins_type='eq_cnt',mth_col=mth_col,base_mth=base_mth,cmp_mth=None,col=None,adj_bin=False,cat_cols=False,path=self.model_path+'/bins_report')
            fea_bins_report=self.cols_bins_rpt[self.cols_bins_rpt.col_name.isin(self.cols)]
            psi_data=fea_bin.psi_mth_avg()
            self.cols_bins_rpt=fea_bins_report
        #系数/重要性、vif、psi、iv、0值率、缺失率
        print("正在计算入模特征效果...")
        fea_report1=pd.merge(left=fea_bins_report[['col_name','iv','ks']].drop_duplicates(),right=psi_data,left_on='col_name',right_on='var')[['col_name','iv','ks','psi']]
        fea_report2=pd.merge(left=self.cols_weight,right=fea_report1,on='col_name')
        fea_report3=pd.merge(left=fea_report2,right=pd.DataFrame({'col_name':self.cols,'missing_rate':self.df[self.cols].isnull().sum()/len(self.df),'0_rate':(self.df[self.cols]==0).mean()}).reset_index(drop=True),on='col_name')
        cs=cols_select(df=self.df,cols=self.cols,target=self.target,vif_thres=np.inf,path=self.model_path)
        cols_copy, filter_details=cs.vif_select()
        fea_report=pd.merge(left=fea_report3,right=filter_details[filter_details.Step==filter_details.Step.min()].rename(columns={'Feature':'col_name','VIF':'vif'})[['col_name','vif']],on='col_name')

        with pd.ExcelWriter(self.model_path+'/'+self.model_type+'_'+str(group)+'bins_full_report.xlsx', engine='openpyxl') as writer:
            confusion_matrix.to_excel(writer,index=None,sheet_name='model_report')
            bins_rpt_all.to_excel(writer,index=None,sheet_name='score_bins_report')
            fea_bins_report.to_excel(writer,index=None,sheet_name='feature_bins_report')
            fea_report.to_excel(writer,index=None,sheet_name='feature_report')
        #画图
        print("正在绘图...")
        #测试集KS
        thres.sort()
        thres_train.sort()
        thres_all.sort()
        plt.figure(figsize=(15,3))
        plt.subplot(131)
        plt.title('TEST_KS', fontsize=15)
        plt.plot(thres, tpr, color='b',label='TPR_CURVE ')
        plt.plot(thres, fpr, color='r',label='FPR_CURVE ')
        plt.plot(thres, tpr-fpr, color='g',linestyle='--',label='TPR-FPR (KS = '+str(round(ks,2))+')')
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('KS',fontsize=12)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        #训练集KS
        plt.subplot(132)
        plt.title('TRAIN_KS', fontsize=15)
        plt.plot(thres_train, tpr_train, color='b',label='TPR_CURVE ')
        plt.plot(thres_train, fpr_train, color='r',label='FPR_CURVE ')
        plt.plot(thres_train, tpr_train-fpr_train, color='g',linestyle='--',label='TPR-FPR (KS = '+str(round(ks_train,2))+')')
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('KS',fontsize=12)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        #全量数据KS
        plt.subplot(133)
        plt.title('ALL_KS', fontsize=15)
        plt.plot(thres_all, tpr_all, color='b',label='TPR_CURVE ')
        plt.plot(thres_all, fpr_all, color='r',label='FPR_CURVE ')
        plt.plot(thres_all, tpr_all-fpr_all, color='g',linestyle='--',label='TPR-FPR (KS = '+str(round(ks_all,2))+')')
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('KS',fontsize=12)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        plt.savefig(self.model_path+'/ks_curve_'+self.model_type+'_'+self.adj+'.png',bbox_inches='tight')
        
        #画图ROC
        #测试集
        plt.figure(figsize=(15,3))
        plt.subplot(131)
        roc_auc=auc(fpr, tpr)
        plt.plot(fpr, tpr, color='blue', label='ROC_CURVE (AUC = {:.4f})'.format(roc_auc))
        plt.plot([0, 1], [0, 1], color='red', linestyle='--')  # 随机猜测的对角线
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('FPR',fontsize=12)
        plt.ylabel('TPR',fontsize=12)
        plt.title('TEST_ROC_AUC', fontsize=15)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        #训练集
        plt.subplot(132)
        roc_auc_train=auc(fpr_train, tpr_train)
        plt.plot(fpr_train, tpr_train, color='blue', label='ROC_CURVE (AUC = {:.4f})'.format(roc_auc_train))
        plt.plot([0, 1], [0, 1], color='red', linestyle='--')  # 随机猜测的对角线
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('FPR',fontsize=12)
        plt.ylabel('TPR',fontsize=12)
        plt.title('TRAIN_ROC_AUC', fontsize=15)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        #全量数据集
        plt.subplot(133)
        roc_auc_all=auc(fpr_all, tpr_all)
        plt.plot(fpr_all, tpr_all, color='blue', label='ROC_CURVE (AUC = {:.4f})'.format(roc_auc_all))
        plt.plot([0, 1], [0, 1], color='red', linestyle='--')  # 随机猜测的对角线
        plt.xlim([0.0, 1.05])
        plt.ylim([0.0, 1.05])
        plt.xlabel('FPR',fontsize=12)
        plt.ylabel('TPR',fontsize=12)
        plt.title('ALL_ROC_AUC', fontsize=15)
        plt.legend(loc='lower right',fontsize=7,framealpha=0.5)
        plt.savefig(self.model_path+'/roc_curve_'+self.model_type+'_'+self.adj+'.png',bbox_inches='tight')
        
        #画图LIFT
        #测试集
        plt.figure(figsize=(15,3))
        plt.subplot(131)
        lift = tpr / (1 - fpr)
        plt.title('TEST_LIFT', fontsize=15)
        plt.plot(thres, lift, color='blue', label='LIFT_CURVE')
        plt.axhline(y=1, color='red', linestyle='--')  # 基准线
        plt.legend(loc='upper left',fontsize=7,framealpha=0.5)
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('LIFT',fontsize=12)
        #训练集
        plt.subplot(132)
        lift_train = tpr_train / (1 - fpr_train)
        plt.title('TRAIN_LIFT', fontsize=15)
        plt.plot(thres_train, lift_train, color='blue', label='LIFT_CURVE')
        plt.axhline(y=1, color='red', linestyle='--')  # 基准线
        plt.legend(loc='upper left',fontsize=7,framealpha=0.5)
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('LIFT',fontsize=12)
        #全量数据集
        plt.subplot(133)
        lift_all = tpr_all / (1 - fpr_all)
        plt.title('ALL_LIFT', fontsize=15)
        plt.plot(thres_all, lift_all, color='blue', label='LIFT_CURVE')
        plt.axhline(y=1, color='red', linestyle='--')  # 基准线
        plt.legend(loc='upper left',fontsize=7,framealpha=0.5)
        plt.xlabel('THRES',fontsize=12)
        plt.ylabel('LIFT',fontsize=12)
        plt.savefig(self.model_path+'/lift_curve_'+self.model_type+'_'+self.adj+'.png',bbox_inches='tight')
        
        #预测概率的分布
        #测试集
        plt.figure(figsize=(15,3))
        plt.subplot(131)
        plt.xlim([0.0, 1.0])
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['y_pred_prob']=self.xtest[self.target],self.xtest['y_pred_prob']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.y_pred_prob,shade=True)
        sbn.kdeplot(card_score_df_bad.y_pred_prob,shade=True,color='r')
        plt.title('TEST_DISTRIBUTION', fontsize=15)
        plt.xlabel('Y_PRED_PROB',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        #训练集
        plt.subplot(132)
        plt.xlim([0.0, 1.0])
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['y_pred_prob']=self.xtrain[self.target],self.xtrain['y_pred_prob']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.y_pred_prob,shade=True)
        sbn.kdeplot(card_score_df_bad.y_pred_prob,shade=True,color='r')
        plt.title('TRAIN_DISTRIBUTION', fontsize=15)
        plt.xlabel('Y_PRED_PROB',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        #全量数据集
        plt.subplot(133)
        plt.xlim([0.0, 1.0])
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['y_pred_prob']=self.df[self.target],self.df['y_pred_prob']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.y_pred_prob,shade=True)
        sbn.kdeplot(card_score_df_bad.y_pred_prob,shade=True,color='r')
        plt.title('ALL_DISTRIBUTION', fontsize=15)
        plt.xlabel('Y_PRED_PROB',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        plt.savefig(self.model_path+'/prob_distribution_'+self.model_type+'_'+self.adj+'.png',bbox_inches='tight')
        plt.show()
        #打分的分布
        plt.figure(figsize=(15,3))
        plt.subplot(131)
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['score']=self.xtest[self.target],self.xtest['score']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.score,shade=True)
        sbn.kdeplot(card_score_df_bad.score,shade=True,color='r')
        plt.title('TEST_DISTRIBUTION', fontsize=15)
        plt.xlabel('SCORE',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        #训练集
        plt.subplot(132)
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['score']=self.xtrain[self.target],self.xtrain['score']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.score,shade=True)
        sbn.kdeplot(card_score_df_bad.score,shade=True,color='r')
        plt.title('TRAIN_DISTRIBUTION', fontsize=15)
        plt.xlabel('SCORE',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        #全量数据集
        plt.subplot(133)
        card_score_df=pd.DataFrame()
        card_score_df['target'],card_score_df['score']=self.df[self.target],self.df['score']
        card_score_df_bad=card_score_df[card_score_df.target==1]
        card_score_df_good=card_score_df[card_score_df.target==0]
        sbn.kdeplot(card_score_df_good.score,shade=True)
        sbn.kdeplot(card_score_df_bad.score,shade=True,color='r')
        plt.title('ALL_DISTRIBUTION', fontsize=15)
        plt.xlabel('SCORE',fontsize=12)
        plt.ylabel('Density',fontsize=12)
        plt.savefig(self.model_path+'/score_distribution_'+self.model_type+'_'+self.adj+'.png',bbox_inches='tight')
        plt.show()
        print("正在调整报告"+self.model_path+'/'+self.model_type+'_'+str(group)+'bins_full_report.xlsx'+"的界面设计样式...")
        print("---------------------------------------------")
        file_path=self.model_path+'/'+self.model_type+'_'+str(group)+'bins_full_report.xlsx'
        # 模型报告样式设计
        wb = load_workbook(file_path)
        ws = wb['model_report']  # 获取指定的Sheet页
        # 定义表头行样式
        deep_blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')  # 深蓝底
        white_bold_font = Font(color='FFFFFF', bold=True)  # 白字加粗
        # 遍历列名所在的单元格，设置样式
        for col in ws.iter_cols(1, ws.max_column):  # 遍历所有列
            header_cell = col[0]  # 列名的单元格（第一行）
            header_cell.fill = deep_blue_fill  # 设置背景色为深蓝色
            header_cell.font = white_bold_font  # 设置字体为白色加粗
        # 去除网格线
        ws.sheet_view.showGridLines = False
        # 加载图片
        path_list=[self.model_path+'/ks_curve_'+self.model_type+'_'+self.adj+'.png',self.model_path+'/roc_curve_'+self.model_type+'_'+self.adj+'.png',self.model_path+'/lift_curve_'+self.model_type+'_'+self.adj+'.png'
                  ,self.model_path+'/prob_distribution_'+self.model_type+'_'+self.adj+'.png',self.model_path+'/score_distribution_'+self.model_type+'_'+self.adj+'.png']
        a=1
        for i in path_list:
            img = Image(i)  # 加载图片
            img.width = 880  # 设置图片宽度（单位：像素）
            img.height = 220  # 设置图片高度（单位：像素）
            ws.add_image(img, 'H'+str(a))  # 将图片插入到Ha单元格
            a=a+13
            wb.save(file_path)
        #分数分箱报告样式设计
        ws = wb['score_bins_report']  # 获取指定的Sheet页
        # 定义不同列的数据条颜色
        column_colors = {'bad_cnt': 'FFFF0000', 'bad_rate': 'FFFF0000','woe': 'FFFF0000', 'good_cnt': '90EE90', 'good_rate': '90EE90','lift': '5B9BD5','ks_bin': '5B9BD5'}
        # 定义色阶规则
        red_white_scale = ColorScaleRule(
            start_type='min',  # 最小值
            start_color=Color(rgb='FFFFFFFF'),  # 白色
            end_type='max',    # 最大值
            end_color=Color(rgb='FFFF0000')  # 红色
        )
        green_white_scale = ColorScaleRule(
            start_type='min',  # 最小值
            start_color=Color(rgb='FFFFFFFF'),  # 白色
            end_type='max',    # 最大值
            end_color=Color(rgb='90EE90')  # 绿色
        )
        #定义色阶列
        column_scale={'ks_grp':red_white_scale, 
                      'auc_grp':green_white_scale}
        # 指定需要添加数据条的列
        columns_to_format = ['bad_cnt', 'bad_rate', 'woe', 'good_cnt','good_rate','lift','ks_bin','ks_grp','auc_grp']  # 需要添加数据条的列名
        # 遍历每一列，添加数据条
        for column_name in columns_to_format:
            # 获取列字母（例如：'Age' -> 'B', 'Score' -> 'C'）
            column_index = bins_rpt_all.columns.get_loc(column_name) + 1  # 列索引（从1开始）
            column_letter = chr(ord('A') + column_index - 1)  # 转换为列字母
            # 确定数据范围（从第2行开始，到数据最后一行）
            start_row = 2  # 数据从第2行开始（第1行是表头）
            end_row = len(bins_rpt_all) + 1  # 数据结束行
            data_range = f'{column_letter}{start_row}:{column_letter}{end_row}'  # 数据范围
            if column_name in list(column_colors.keys()):
                # 为当前列定义数据条规则
                data_bar_rule = DataBarRule(
                    start_type='min',  # 数据条起始值为数值
                    start_value=bins_rpt_all[column_name].min(),
                    end_value=bins_rpt_all[column_name].max(),
                    end_type='max',    # 数据条结束值为最大值
                    color=Color(rgb=column_colors[column_name]),  # 使用列对应的颜色
                    showValue=True,     # 显示数值
                    minLength=None,    # 最小长度
                    maxLength=None     # 最大长度
                )
                # 为当前列添加数据条
                ws.conditional_formatting.add(data_range, data_bar_rule)
            else:
                ws.conditional_formatting.add(data_range, column_scale[column_name])
        # 定义表头样式 遍历列名所在的单元格，设置样式
        for col in ws.iter_cols(1, ws.max_column):  # 遍历所有列
            header_cell = col[0]  # 列名的单元格（第一行）
            header_cell.fill = deep_blue_fill  # 设置背景色为深蓝色
            header_cell.font = white_bold_font  # 设置字体为白色加粗
        # 去除网格线
        ws.sheet_view.showGridLines = False
        # 保存修改后的Excel文件
        wb.save(file_path)
        #特征分箱报告样式
        ws = wb['feature_bins_report']  # 获取指定的Sheet页
        # 定义不同列的数据条颜色
        #定义色阶列
        column_scale={'ks':red_white_scale, 
                      'iv':green_white_scale}
        # 指定需要添加数据条的列
        columns_to_format = ['bad_cnt', 'bad_rate', 'woe', 'good_cnt','good_rate','lift','ks_bin','ks','iv']  # 需要添加数据条的列名
        # 遍历每一列，添加数据条
        for column_name in columns_to_format:
            # 获取列字母（例如：'Age' -> 'B', 'Score' -> 'C'）
            column_index = fea_bins_report.columns.get_loc(column_name) + 1  # 列索引（从1开始）
            column_letter = chr(ord('A') + column_index - 1)  # 转换为列字母
            # 确定数据范围（从第2行开始，到数据最后一行）
            start_row = 2  # 数据从第2行开始（第1行是表头）
            end_row = len(fea_bins_report) + 1  # 数据结束行
            data_range = f'{column_letter}{start_row}:{column_letter}{end_row}'  # 数据范围
            if column_name in list(column_colors.keys()):
                # 为当前列定义数据条规则
                data_bar_rule = DataBarRule(
                    start_type='min',  # 数据条起始值为数值
                    start_value=fea_bins_report[column_name].min(),
                    end_value=fea_bins_report[column_name].max(),
                    end_type='max',    # 数据条结束值为最大值
                    color=Color(rgb=column_colors[column_name]),  # 使用列对应的颜色
                    showValue=True,     # 显示数值
                    minLength=None,    # 最小长度
                    maxLength=None     # 最大长度
                )
                # 为当前列添加数据条
                ws.conditional_formatting.add(data_range, data_bar_rule)
            else:
                ws.conditional_formatting.add(data_range, column_scale[column_name])
        # 定义表头样式 遍历列名所在的单元格，设置样式
        for col in ws.iter_cols(1, ws.max_column):  # 遍历所有列
            header_cell = col[0]  # 列名的单元格（第一行）
            header_cell.fill = deep_blue_fill  # 设置背景色为深蓝色
            header_cell.font = white_bold_font  # 设置字体为白色加粗
        # 去除网格线
        ws.sheet_view.showGridLines = False
        # 保存修改后的Excel文件
        wb.save(file_path)
        #特征效果报告样式
        ws=wb['feature_report']
        # 定义不同列的数据条颜色
        column_colors = {
            'missing_rate': '5B9BD5',  # 红色
            '0_rate': '5B9BD5'  # 红色
        }
        #定义色阶列
        column_scale={'ks':red_white_scale, 'vif':red_white_scale,'iv':green_white_scale,'psi':green_white_scale}
        # 指定需要添加数据条的列
        columns_to_format = ['missing_rate', '0_rate','ks','vif','iv','psi']  # 需要添加数据条的列名
        # 遍历每一列，添加数据条
        for column_name in columns_to_format:
            # 获取列字母（例如：'Age' -> 'B', 'Score' -> 'C'）
            column_index = fea_report.columns.get_loc(column_name) + 1  # 列索引（从1开始）
            column_letter = chr(ord('A') + column_index - 1)  # 转换为列字母
            # 确定数据范围（从第2行开始，到数据最后一行）
            start_row = 2  # 数据从第2行开始（第1行是表头）
            end_row = len(fea_report) + 1  # 数据结束行
            data_range = f'{column_letter}{start_row}:{column_letter}{end_row}'  # 数据范围
            if column_name in list(column_colors.keys()):
                # 为当前列定义数据条规则
                data_bar_rule = DataBarRule(
                    start_type='min',  # 数据条起始值为数值
                    start_value=fea_report[column_name].min(),
                    end_value=fea_report[column_name].max(),
                    end_type='max',    # 数据条结束值为最大值
                    color=Color(rgb=column_colors[column_name]),  # 使用列对应的颜色
                    showValue=True,     # 显示数值
                    minLength=None,    # 最小长度
                    maxLength=None     # 最大长度
                )
                # 为当前列添加数据条
                ws.conditional_formatting.add(data_range, data_bar_rule)
            else:
                ws.conditional_formatting.add(data_range, column_scale[column_name])
        # 定义表头样式 遍历列名所在的单元格，设置样式 遍历列名所在的单元格，设置样式
        for col in ws.iter_cols(1, ws.max_column):  # 遍历所有列
            header_cell = col[0]  # 列名的单元格（第一行）
            header_cell.fill = deep_blue_fill  # 设置背景色为深蓝色
            header_cell.font = white_bold_font  # 设置字体为白色加粗
        # 去除网格线
        ws.sheet_view.showGridLines = False
        # 保存修改后的Excel文件
        wb.save(file_path)
        return bins_rpt_all,confusion_matrix,fea_bins_report,fea_report





class uns_model:
    '''
    无监督学习，现包含主成分分析和kmeans聚类
    mean_key：主键list
    target：目标变量str
    df：输入数据集
    cols：输入特征list
    
    用法
    um=uns_model(df.fillna(0),cols,target='target',mean_key='id',path='step6_uns_model')
    pca_cols,pca_data,pca_detail=um.pca() 主成分分析
    um.get_keams_cluster() 获取聚类个数
    data,kmeans_detail=um.kmeans(n_clusters=10)聚类
    '''
    
    def __init__(self,df,cols,target=None,mean_key=[],path='step6_uns_model'):
        self.df=df
        self.df_copy=df.copy() 
        self.cols=cols
        self.mean_key=mean_key
        self.target=target
        
        self.path=path
        dir_name=path
        current_path = os.getcwd()
        path=os.path.join(os.getcwd(), dir_name)
        try:
            os.mkdir(path)  # 使用 os.makedirs(path, exist_ok=True) 可以避免多级目录的创建
            print(f"Directory '{dir_name}' created successfully at '{current_path}'.")
        except FileExistsError:
            print(f"Directory '{dir_name}' already exists at '{current_path}'.")
        except PermissionError:
            print(f"Permission denied: Unable to create directory at '{current_path}'.")


    
    #主成分分析，必要时只需变更累计方差解释比例var_ratio_cumsum阈值即可，默认为80%
    def pca(self,var_ratio_cumsum=0.8):
        print("---------------------------------------------")
        data=self.df.copy()
        t=0
        variance_ratio_sum=0
        while variance_ratio_sum<=var_ratio_cumsum:
            t=t+1
            pca=PCA(t)
            pca.fit(data[self.cols])
            principal_components = pca.components_
            # 获取每个主成分解释的方差比例
            explained_variance_ratio = pca.explained_variance_ratio_
            # 对数据进行降维
            reduced_data = pca.transform(data[self.cols])
            # 获取主成分得分
            principal_scores = reduced_data
            # 获取累积方差解释比例
            variance_ratio_sum=sum(explained_variance_ratio)
            cumulative_variance_ratio = np.cumsum(explained_variance_ratio)
            weighted_scores = principal_scores * explained_variance_ratio
            #final_scores = weighted_scores.sum(axis=1)
            # 打印结果
        print('累积方差解释比例:',variance_ratio_sum)
        print('主成分个数:',len(weighted_scores[0]))
        pca_cols=[]
        for i in range(len(weighted_scores[0])):
            pca_cols.append('F'+str(i+1))
        pca_data=pd.DataFrame(weighted_scores,columns=pca_cols)
        pca_data=pd.concat([data[self.mean_key],pca_data,data[self.target]],axis=1)
        pca_detail=pd.DataFrame(np.round(pca.components_,4),columns=self.cols).set_index(pd.Index(pca_cols))
        print('各主成分相关性：')
        print(pca_data[pca_cols].corr())
        joblib.dump(pca_cols,self.path+'/pca_cols.pkl')
        pca_data.to_csv(self.path+'/pca_data.csv',index=None)
        joblib.dump(pca,self.path+'/pca_model.pkl')
        pca_detail.to_csv(self.path+'/pca_details.csv',index=None)
        pca_data[pca_cols].corr().to_excel(self.path+'/pca_cols_corr.xlsx',index=None)
        print("---------------------------------------------")
        return pca_cols,pca_data,pca_detail


    def get_keams_cluster(self):
        print("---------------------------------------------")
        data=self.df.copy()
        # 计算不同聚类数下的损失值
        sse = []
        silhouette_scores = []
        for k in range(2, 30):
            kmeans = KMeans(n_clusters=k, random_state=666)
            cluster_labels = kmeans.fit_predict(data[self.cols])
            silhouette_avg = silhouette_score(data[self.cols], cluster_labels)
            silhouette_scores.append(silhouette_avg)
            kmeans.fit(data[self.cols])
            sse.append(kmeans.inertia_)
        
        # 绘制肘部法曲线
        plt.figure(figsize=(8, 12))
        plt.subplot(211)
        plt.plot(range(2, 30), sse, marker='o')
        plt.xlabel('Number of clusters')
        plt.ylabel('SSE')
        plt.title('Elbow Method for Optimal K')
        plt.xticks(range(2, 30))
        plt.grid(True)
        plt.savefig(self.path+'/SSE肘部法评估.png',bbox_inches='tight')
        # 绘制轮廓系数曲线
        plt.figure(figsize=(8, 12))
        plt.subplot(212)
        plt.plot(range(2, 30), silhouette_scores, marker='o')
        plt.xlabel('Number of clusters')
        plt.ylabel('Silhouette Score')
        plt.title('Silhouette Score for Optimal K')
        plt.xticks(range(2, 30))
        plt.grid(True)
        plt.savefig(self.path+'/轮廓系数评估.png',bbox_inches='tight')
        plt.show()
        print("---------------------------------------------")

    def kmeans(self,n_clusters):
        print("---------------------------------------------")
        data=self.df.copy()
        kmeans = KMeans(n_clusters=n_clusters, random_state=666)
        cluster_labels = kmeans.fit_predict(data[self.cols])
        data['cluster_labels']=cluster_labels
        print('列为每个变量的均值，行为各均值对应的聚类类别')
        kmeans_detail=pd.DataFrame(kmeans.cluster_centers_,columns=self.cols).set_index(pd.Index(np.unique(cluster_labels).tolist()))
        print(kmeans_detail)
        joblib.dump(cluster_labels,self.path+'/cluster_labels.pkl')
        kmeans_detail.to_excel(self.path+'/kmeans_center.xlsx',index=None)
        print("---------------------------------------------")
        return data,kmeans_detail


class linear_model:
    """
    Paras:
    df:输入数据集
    cols:解释变量list
    target：被解释变量str
    iv_target：将要被工具变量回归模拟的列
    iv_col：工具变量列
    
    用法：
    lm=linear_model(df,cols,target='target',iv_target='col2',iv_col='col1',path='step6_linear_model')
    ols=lm.ols() 当被解释变量为二分类时，自动选择逻辑回归logit作为模型
    ols1,ols2=lm.IV()
    """
    def __init__(self,df,cols,iv_target,iv_col,target,path='step6_linear_model'):
        self.df=df
        self.df_copy=df.copy()
        self.cols=cols
        self.target=target
        self.iv_target=iv_target
        self.iv_col=iv_col
        
        self.path=path
        dir_name=path
        current_path = os.getcwd()
        path=os.path.join(os.getcwd(), dir_name)
        try:
            os.mkdir(path)  # 使用 os.makedirs(path, exist_ok=True) 可以避免多级目录的创建
            print(f"Directory '{dir_name}' created successfully at '{current_path}'.")
        except FileExistsError:
            print(f"Directory '{dir_name}' already exists at '{current_path}'.")
        except PermissionError:
            print(f"Permission denied: Unable to create directory at '{current_path}'.")

    
    def ols(self):
        print("---------------------------------------------")
        # 添加显著性标注
        def significance_mark(p_value):
            if p_value < 0.01:
                return '***'
            elif p_value < 0.05:
                return '**'
            elif p_value < 0.1:
                return '*'
            else:
                return ''
        
        if len(self.df[self.target].unique())==2:#判断是否使用逻辑回归
            model = smf.logit(self.target+' ~ '+' + '.join(self.cols), data=self.df).fit()
            self.model_name='logit'
            summary = model.summary2().tables[1]
            summary['Significance'] = summary['P>|z|'].apply(significance_mark)
        else:
            model = smf.ols(self.target+' ~ '+' + '.join(self.cols), data=self.df).fit()
            self.model_name='ols'
            model.summary2().tables[2].to_excel(self.path+'/ols_model_normal_distribution_estimate.xlsx',header=False,index=None)
            print("""
Durbin-Watson：用于检测回归残差的自相关性。值在 0 到 4 之间，2 表示没有自相关，值小于 2 表示正自相关，大于 2 表示负自相关。
Omnibus：用于检验残差是否符合正态分布的统计量。值越大，拒绝正态分布的假设的可能性越高。
Prob(Omnibus)：Omnibus 测试的 p 值。若 p 值小于显著性水平（通常为 0.05），则拒绝残差符合正态分布的假设。
Skew：残差的偏度，反映分布的对称性。值为 0 表示对称，正值表示右偏，负值表示左偏。
Kurtosis：残差的峰度，反映分布的尖锐程度。正态分布的峰度为 3，值大于 3 表示分布比正态分布更尖锐，值小于 3 表示更平坦。
Cond. No.（条件数）：用于评估模型的多重共线性。如果条件数大于 30，可能存在多重共线性的问题。
Jarque-Bera : 是一种用于检验数据样本是否来自正态分布的统计检验。它基于数据的偏度（Skewness）和峰度（Kurtosis）计算得出。
Prob(JB)： 是与 Jarque-Bera 统计量相关的 p 值。它表示在零假设（即残差来自正态分布）的条件下，观察到的 Jarque-Bera 统计量值或更极端值的概率。该p值越大（至少大于0.05）说明残差越符合正态分布。
            """)
            # 创建显著性标注列
            summary = model.summary2().tables[1]
            summary['Significance'] = summary['P>|t|'].apply(significance_mark)
        print(model.summary())
        self.model=model
        model.summary2().tables[0].to_excel(self.path+'/'+self.model_name+'_model_base_estimate.xlsx',header=False,index=None)
        summary.to_excel(self.path+'/'+self.model_name+'_Significance_summary.xlsx')
        joblib.dump(model,self.path+'/'+self.model_name+'_model.pkl')
        print("---------------------------------------------")
        return model

    
    def IV(self):
        print("---------------------------------------------")
        ols1 = smf.ols(self.iv_target+' ~ '+ self.iv_col, data=self.df_copy).fit()
        self.df_copy['ols1']=ols1.predict()
        ols2 = smf.ols(self.target+' ~ '+ 'ols1', data=self.df_copy).fit()
        print("第一阶段回归结果：")
        print(ols1.summary())
        print("\n第二阶段回归结果：")
        print(ols2.summary())
        joblib.dump(ols1,self.path+'/iv_ols1_model.pkl')
        joblib.dump(ols2,self.path+'/iv_ols2_model.pkl')
        print("---------------------------------------------")
        return ols1,ols2
        
